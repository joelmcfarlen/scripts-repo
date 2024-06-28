import argparse
import json
import sys
import os
import boto3
import time
from openpyxl import Workbook
from openpyxl import load_workbook

marker = None

parser = argparse.ArgumentParser(
        description='Creates a list all RDS Instances.')

parser.add_argument('-p',
					'--profiles',
					required=False,
					type=str,
					nargs='+',
					help='The list of AWS profiles to access.')

parser.add_argument('-a',
					'--account_name',
					type=str,
					nargs='+',
					required=False,
					help='The friendly name of a single AWS Account that you are authenticated with.')

parser.add_argument('-r',
					'--regions',
					required=True,
					type=str,
 					nargs='+',
					help='The list of regions to report. REQUIRED')

args = parser.parse_args()

if not args.regions:
	sys.exit('Please provide at least 1 region')

if (not args.account_name) and (not args.profiles) or (args.account_name) and (args.profiles):
	sys.exit('Use either profiles or one account name, not both.')

if args.profiles:
	profiles = args.profiles

if args.account_name:
	profiles = args.account_name

regions = args.regions

InstanceId = 'i-08940bb036c8678f2'
sts_client = boto3.client('sts','us-east-1')

account = sts_client.get_caller_identity().get('Account')
for profile in profiles:
    print(profile)
    if args.profiles:
        boto3.setup_default_session(profile_name=profile)
    account_name = profile
    sts_client = boto3.client('sts','us-east-1')
    account = sts_client.get_caller_identity().get('Account')
    workbook_name = f"SSM_Instance_Inventory.xlsx"
    headers_row =  [
        'AccountNumber','AccountName','Region','InstanceId','ApplicationType',
        'Architecture','InstalledTime','Name','PackageId',
        'Publisher','Release','Summary','URL','Version'
    ]
    try: 
        workbook = load_workbook(filename = f"./output/{workbook_name}")
        if account_name in workbook.sheetnames:
            worksheet = workbook[account_name]
        else:
            workbook.create_sheet(account_name)
            worksheet = workbook[account_name]
            worksheet.append(headers_row)
    except FileNotFoundError:
        print(f"File not found. Creating one.")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = account_name
        worksheet.append(headers_row)

    for region in regions:
        print(f'  {region}')
        try:
            ssm_client = boto3.client('ssm',region)
            next_token = None
            first_iteration = False
            if next_token == None:
                inventory = ssm_client.list_inventory_entries(
                     InstanceId=InstanceId,
                     TypeName='AWS:Application'
                )
                inventory_entry = inventory['Entries']
                for entry in inventory_entry:
                    ApplicationType = entry.get('ApplicationType','')
                    Architecture = entry.get('Architecture','')
                    InstalledTime = entry.get('InstalledTime','')
                    Name = entry.get('Name','')
                    PackageId = entry.get('PackageId','')
                    Publisher = entry.get('Publisher','')
                    Release = entry.get('Release','')
                    Summary = entry.get('Summary','')
                    Url = entry.get('Url','')
                    Version = entry.get('Version','')
                    new_row = [
                        account,account_name,region,InstanceId,ApplicationType,
                        Architecture,InstalledTime,Name,PackageId,Publisher,
                        Release,Summary,Url,Version
                    ]
                    worksheet.append(new_row)  
                    next_token  = inventory.get('NextToken', None)
            while next_token is not None:                
                if first_iteration != False:
                    inventory = ssm_client.list_inventory_entries(
                         InstanceId=InstanceId,
                         TypeName='AWS:Application',
                         NextToken=next_token
                    )
                first_iteration = True
                inventory_entry = inventory['Entries']
                for entry in inventory_entry:
                    ApplicationType = entry.get('ApplicationType','')
                    Architecture = entry.get('Architecture','')
                    InstalledTime = entry.get('InstalledTime','')
                    Name = entry.get('Name','')
                    PackageId = entry.get('PackageId','')
                    Publisher = entry.get('Publisher','')
                    Release = entry.get('Release','')
                    Summary = entry.get('Summary','')
                    Url = entry.get('Url','')
                    Version = entry.get('Version','')
                    new_row = [
                        account,account_name,region,InstanceId,ApplicationType,
                        Architecture,InstalledTime,Name,PackageId,Publisher,
                        Release,Summary,Url,Version
                    ]
                    worksheet.append(new_row)                                                  
                try:
                    next_token = inventory.get('NextToken')
                except:
                    next_token = None
            worksheet.auto_filter.ref = worksheet.dimensions
            os.makedirs(f"./output/", exist_ok=True)
            workbook.save(f"./output/{workbook_name}")
        except Exception as e:
            print(e)